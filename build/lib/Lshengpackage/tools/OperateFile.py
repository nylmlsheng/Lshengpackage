# -*- coding: UTF-8 -*-
import os
from collections import defaultdict
from openpyxl import Workbook, load_workbook
import pandas as pd


# 删除文件夹下所有文件
def del_Files(dir_path):
    """
    :param dir_path: 文件夹路径
    :type dir_path:  str
    :return:
    :rtype:
    """
    if os.path.isfile(dir_path):
        try:
            os.remove(dir_path)  # 这个可以删除单个文件，不能删除文件夹
        except BaseException as e:
            return e
    elif os.path.isdir(dir_path):
        file_lis = os.listdir(dir_path)
        for file_name in file_lis:
            # if file_name != 'wibot.log':
            tf = os.path.join(dir_path, file_name)
            del_Files(tf)
        return True


# 获取文件夹下所有文件名
def file_in_Fol(file_path):
    file_name_list = defaultdict(list)
    """
    :param file_path:  文件夹路径
    :type file_path:  str
    :return: list
    :rtype:
    """
    filenames = os.listdir(file_path)
    for i in filenames:
        scr_name = i.split('.')[0]
        _type = i.split('.')[-1]
        file_name_list[_type].append(scr_name)
    return file_name_list


# 文件夹不存在，创建文件夹
def mk_Fol(path):
    """
    检测路径是否存在，不存在则创建路径
    """
    pa = os.path.exists(path)
    if pa is not False:
        pass
    else:
        os.mkdir(path=path)
        return True


# xls_转_xlsx程序，文件夹为程序exc下，转换后自动读取
def xls_To_xlsx(path, f_name, sheet_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    path 文件夹路径
    f 读取文件的文件名
    """
    file_name_be, suff = os.path.splitext(f_name)  # 路径进行分割，分别为文件路径和文件后缀
    # print(file_name_be, suff)
    if suff == '.xls':
        # ui.printf('将对{}文件进行转换...'.format(f_name))
        # print(path + '/' + f)
        data = pd.DataFrame(pd.read_excel(path + '/' + f_name, sheet_name=sheet_name))  # 读取xls文件
        data.to_excel(os.getcwd() + '/exc/' + file_name_be + '.xlsx', sheet_name=sheet_name, index=False)  # 格式转换
        return file_name_be + '.xlsx'
    else:
        pass


def copy_sheet(src_xlsx, ssheetname, dst_xlsx, nsheetname=None):
    if nsheetname == None:
        nsheetname = ssheetname
    try:
        sw = load_workbook(f'{src_xlsx}')
    except KeyError:
        raise KeyError('旧工作簿不存在 The old xlsx is not exists')

    try:
        dw = load_workbook(f'{dst_xlsx}')
    except FileNotFoundError:
        dw = Workbook()

    try:
        sheet = dw[f'{nsheetname}']
    except KeyError:
        sheet = dw.create_sheet(f'{nsheetname}')

    try:
        src_sheet = sw[f'{ssheetname}']
    except KeyError:
        raise KeyError('源工作簿文件不存在该工作簿 The sheet does not exist in the source file')

    # for i in range(1):
    #     sheet.delete_rows(2000)
    for sheet in dw:
        sheet.delete_rows(1, 3000)

    for row in src_sheet.iter_rows():
        # print(row)
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        sheet.append(row_list)
    dw.save(f'{dst_xlsx}')
